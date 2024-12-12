import openpyxl

__version__ = '0.0.1'

class ExcelLib():

    def __init__(self):
        ROBOT_LIBRARY_SCOPE = 'GLOBAL'
        ROBOT_LIBRARY_VERSION = __version__
        self.wb = None
        self.xl = None
        self.filename = None

    def Add_Workbook(self,ExcelPath):
        """
        Add Excel Workbook

        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Add Workbook           |  C:\\Python_Work\\SampleTest.xlsx |

        """
        self.xl = openpyxl.Workbook()
        self.xl.save(ExcelPath)
    
    def Add_Worksheet(self,SheetName):
        """
        Add Excel Worksheet

        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Add Worksheet           |  SheetName |
        
        """
        self.wb.create_sheet(SheetName)
        
    def Set_Font_Bold(self, SheetName,iRow,iCol,Status):
        """
        Set Excel Font Bold Style

        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Set Font Bold        |  SheetName  | 1  | 1  |True

        """
        ws = self.wb[SheetName]
        cell = ws.cell(row=int(iRow),column=int(iCol))
        cell.font = cell.font.copy(bold=Status)
        
    def Set_Font_Color(self, SheetName,iRow,iCol,ColorCode):
        """
        Set Excel Font Color Style

        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Set Font Color        |  SheetName  | 1  | 1  | 3

        """
        ws = self.wb[SheetName]
        cell = ws.cell(row=int(iRow),column=int(iCol))
        cell.font = cell.font.copy(color=ColorCode)

    
    def Insert_Row(self, SheetName,iRow):
       """
       Insert Excel Row

       Example:

       | *Keywords*           |  *Parameters*                                      |
       | Insert Row          |  SheetName  | Row  |

       """
       ws = self.wb[SheetName]
       ws.insert_rows(int(iRow))

    def Delete_Row(self, SheetName,iRow):
       """
       Delete Excel Row

       Example:

       | *Keywords*           |  *Parameters*                                      |
       | Insert Row          |  SheetName  | Row  |

       """
       ws = self.wb[SheetName]
       ws.delete_rows(int(iRow))

    def Open_Excel(self, ExcelPath):
        """
        Opens the Excel file to read from the path provided in the file path parameter.

        Arguments:
                |  File Path (string) | The Excel file name or path will be opened.  |
        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Open Excel           |  C:\\Python_Work\\SampleTest.xlsx  |

        """
        self.wb = openpyxl.load_workbook(ExcelPath,data_only=True)
        self.filename = ExcelPath

    def Open_Excel_FORMULA(self, ExcelPath):
        """
        Opens the Excel file to read from the path provided in the file path parameter.

        Arguments:
                |  File Path (string) | The Excel file name or path will be opened.  |
        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Open_Excel_FORMULA          |  C:\\Python_Work\\SampleTest.xlsx  |

        """
        self.wb = openpyxl.load_workbook(ExcelPath,keep_vba=True,data_only=False)
        self.filename = ExcelPath

    def Get_Sheets_Count(self):
        """
        Returns the number of worksheets in the current workbook.

        Example:

        | *Keywords*              |  *Parameters*                                      |
        | Open Excel              |  C:\\Python_Work\\SampleTest.xlsx  |
        | ${sheetcount}    |  Get Sheets Count                                              |

        """
        return len(self.wb.sheetnames)

    def Get_Sheets_Name(self):
        """
        Returns the names of all the worksheets in the current workbook.

        Example:

        | *Keywords*              |  *Parameters*                                      |
        | Open Excel              |  C:\\Python_Work\\SampleTest.xlsx  |
        | ${sheetname}    |  Get Sheets Name                                                    |

        """
        return self.wb.sheetnames

    def Get_Sheets_Name_By_Index(self,Index):
        """
        Returns the names of all the worksheets in the current workbook.

        Example:

        | *Keywords*              |  *Parameters*                                      |
        | Open Excel              |  C:\\Python_Work\\SampleTest.xlsx  |
        | ${sheetname}    |  Get Sheets Name By Index | 1                                                  |

        """
        return  self.wb.worksheets[int(Index)].title

    def Get_Row_Count(self,SheetName):
        """
        Returns the specific number of rows of the sheet name specified.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the row count will be returned from. |
        Example:

        | *Keywords*          |  *Parameters*                                      |
        | Open Excel          |  C:\\Python_Work\\SampleTest.xlsx                   |
        | ${RowCount}           |  Get Row Count                                     | TestSheet1 |

        """
        ws = self.wb[SheetName]
        return ws.max_row


    def Get_Column_Count(self,SheetName):
        """
        Returns the specific number of Column of the sheet name specified.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the row count will be returned from. |
        Example:

        | *Keywords*          |  *Parameters*                                      |
        | Open Excel          |  C:\\Python_Work\\SampleTest.xlsx  |                |
        | ${ColCount}        |  Get Column Count                                        | TestSheet1 |

        """
        ws = self.wb[SheetName]
        return ws.max_column

    def Read_Cell_Data(self,SheetName,iRow,iCol):
        """
        Uses the column and row to return the data from that cell.

        Arguments:
                |  Sheet Name (string)                      | The selected sheet that the cell value will be returned from.             |
                |  Row (int)                                | The selected row that will be returned from.                   |
                |  Column (int)                             | The selected column that will be returned from.                |
        Example:

        | *Keywords*                |  *Parameters*                                             |
        | Open Excel                |  C:\\Python_Work\\SampleTest.xlsx  |      |
        | ${data}    |  Read Cell Data                                        |  Sheet1  |   1  |   1  |

        """
        ws = self.wb[SheetName]
        return  ws.cell(row=int(iRow),column=int(iCol)).value

    def Read_Cell_Data_By_Name(self,SheetName,CellName):
        """
        Uses the cell name to return the data from that cell.

        Arguments:
                |  Sheet Name (string)                      | The selected sheet that the cell value will be returned from.             |
                |  Cell Name (string)                       | The selected cell name that the value will be returned from.              |
        Example:

        | *Keywords*                |  *Parameters*                                             |
        | Open Excel                |  C:\\Python_Work\\SampleTest.xlsx  |      |
        | ${data}    |  Read Cell Data By Name                                      |  Sheet1  |   A1  |     |

        """
        ws = self.wb[SheetName]
        return ws[CellName].value

    def Write_Cell_Data(self,SheetName,iRow,iCol,InputData):
        """
        Write data to cell by using the column and row.

        Arguments:
                |  Sheet Name (string)                      | The selected sheet that the cell will be modified from.                       |
                |  Row (int)                                | The selected row that will be used to modify from.                   |
                |  Column (int)                             | The selected column that will be used to modify from.                |
                |  Value (string)   | Raw value or string value    |
        Example:

        | *Keywords*            |  *Parameters*                                                                     |
        | Open Excel            |  C:\\Python_Work\\SampleTest.xlsx  |                      |       |
        | Write Cell Data |  Sheet1                                        |  1  |    1    |  SampleData     |

        """
        print("Data :"+InputData)
        ws = self.wb[SheetName]
        ws.cell(row=int(iRow),column=int(iCol)).value = str(InputData)

    def Write_Cell_Data_By_Name(self,SheetName,CellName,InputData):
        """
        Write data to cell by using the given sheet name and the given cell that defines by name.

        Arguments:
                |  Sheet Name (string)                      | The selected sheet that the cell will be modified from.                       |
                |  Cell Name (string)                       | The selected cell name that will be used to modified from.                  |
                |  Value (string)   | Raw value or string value    |
        Example:

        | *Keywords*            |  *Parameters*                                                                     |
        | Open Excel            |  C:\\Python_Work\\SampleTest.xlsx  |                      |       |
        | Write Cell Data By Name |  Sheet1                                        |  A2  |  SampleData           |       |

        """
        print("Data :" + InputData)
        ws = self.wb[SheetName]
        ws[CellName].value = str(InputData)

    def Clear_Cell_Data(self,SheetName,iRow,iCol):
        """
        Delete cell data by using the column and row.

        Arguments:
                |  Sheet Name (string)                      | The selected sheet that the cell will be clear value.                       |
                |  Row (int)                                | The selected row that will be used to clear value.                   |
                |  Column (int)                             | The selected column that will be used to clear value.                |
        Example:

        | *Keywords*            |  *Parameters*                                                                     |
        | Open Excel            |  C:\\Python_Work\\SampleTest.xlsx  |                      |       |
        | Clear Cell Data |  Sheet1                                        |  1  |    1    |       |

        """
        ws = self.wb[SheetName]
        ws.cell(row=int(iRow),column=int(iCol)).value = ''

    def Clear_Cell_Data_By_Name(self,SheetName,CellName):
        """
        Delete cell data by using the given sheet name and the given cell that defines by name.

        Arguments:
                |  Sheet Name (string)                      | The selected sheet that the cell will be clear value.                       |
                |  Cell Name (string)                       | The selected cell that will be used to clear value.                  |
        Example:

        | *Keywords*            |  *Parameters*                                                                     |
        | Open Excel            |  C:\\Python_Work\\SampleTest.xlsx  |                      |       |
        | Clear Cell Data By Name |  Sheet1                                        |  A1  |        |       |

        """
        ws = self.wb[SheetName]
        ws[CellName].value = ''

    def Save_Excel(self):
        """
        Saves the Excel file that was opened to write before.

        Example:

        | *Keywords*            |  *Parameters*                                      |
        | Open Excel            |  C:\\Python_Work\\SampleTest.xlsx  |                  |
        | Write Cell Data       |  TestSheet1                                        |  Sheet1  |  1    |  1  |  SampleData  |
        | Save Excel            |                                                    |                  |

        """
        self.wb.save(self.filename)

    def Close_Excel(self):
        """
        Close the Excel file

        Example:

        | *Keywords*            |  *Parameters*                                      |
        | Open Excel            |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |                  |
        | Write Cell Data       |  TestSheet1                                        |  Sheet1  |  1    |  1  |  SampleData  |
        | Close Excel          |    |                                                |                  |

        """
        self.wb.close()
        self.xl = None
